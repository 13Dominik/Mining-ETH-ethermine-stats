from datetime import datetime, timedelta, date
from typing import List, Dict
import requests
import json
import os.path

from pycoingecko import CoinGeckoAPI
import openpyxl
from openpyxl.utils import get_column_letter

from setUp import miner_address

cg = CoinGeckoAPI()

Payout = Dict[date, float]
List_Payouts = List[Payout]


class Miner:
    base_link = 'https://api.ethermine.org/miner/'
    price_eth = cg.get_price(ids='ethereum', vs_currencies='usd')['ethereum']['usd']

    def __init__(self, miner_address: str, *args, **kwargs):
        super(Miner, self).__init__(*args, **kwargs)
        self.miner_address = miner_address

    def get_unpaid_eth(self) -> float:
        """ Returns amount of unpaid_eth [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/dashboard")
        file = json.loads(res.text)
        return round(int(file['data']['currentStatistics']['unpaid']) / 10e17, 4)

    def get_sum_payouts(self) -> float:
        """ Returns sum of payouts [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/payouts")
        file = json.loads(res.text)
        return round(sum([int(elem['amount']) / 10e17 for elem in file['data']]), 4)

    def get_daily_eth(self) -> float:
        """ Returns daily mined eth [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/currentStats")
        file = json.loads(res.text)
        return round(float(file['data']['coinsPerMin']) * 24 * 60, 5)

    def get_min_payment_threshold(self) -> float:
        """ Returns minimum payment threshold [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/settings")
        file = json.loads(res.text)
        return file['data']['minPayout'] / 10e17

    def days_to_next_payout(self) -> int:
        """ Returns days to next payout [Days] """

        return int((self.get_min_payment_threshold() - self.get_unpaid_eth()) / (self.get_daily_eth()))

    def date_of_next_payout(self) -> date:
        """ Returns date of  next payout [Y-M-D] """

        d = datetime.now() + timedelta(days=self.days_to_next_payout())
        return date(d.year, d.month, d.day)

    def get_list_of_payouts(self, order: str = "oldest") -> List_Payouts:
        """
        Returns a list of payout in form List[Dict[date, float]]
        if order == "oldest" list of payouts is sorted from first to last
        else from last to first
        :param order: str
        :return: List_Payouts
        """

        res = requests.get(self.base_link + self.miner_address + "/payouts")
        file = json.loads(res.text)
        list_of_payouts = []

        for payout in file['data']:
            d = datetime.fromtimestamp(payout['paidOn'])
            list_of_payouts.append({date(d.year, d.month, d.day): round(int(payout['amount']) / 10e17, 4)})
        if order == "oldest":
            list_of_payouts.reverse()
        return list_of_payouts

    def get_current_hashrate(self) -> float:
        """
        Returns a current average hashrate of miner [MH/s]
        :return: float
        """
        res = requests.get(self.base_link + self.miner_address + "/dashboard")
        file = json.loads(res.text)
        return round(float(file['data']['statistics'][0]['reportedHashrate']) / 1000000, 1)

    def save_todays_data_to_xlsx(self) -> None:
        """
        If today's data didn't save -save it to xlsx file
        :return: None
        """
        de = Data_Excel()
        if de.checks_if_todays_data_is_saved():
            return
        wb = openpyxl.load_workbook('mining_data.xlsx')
        sheet = wb.active
        column_to_write = get_column_letter(sheet.max_column + 1)
        sheet[column_to_write + "1"] = date(datetime.today().year, datetime.today().month, datetime.today().day)
        sheet[column_to_write + "2"] = self.get_current_hashrate()
        sheet[column_to_write + "3"] = self.get_unpaid_eth()
        sheet[column_to_write + "4"] = self.get_daily_eth()
        sheet[column_to_write + "5"] = self.get_sum_payouts()
        sheet[column_to_write + "6"] = self.days_to_next_payout()
        wb.save('mining_data.xlsx')

        wb = openpyxl.load_workbook('mining_data.xlsx')  # stretching cells
        de.stretch_cells(wb.active)
        wb.save('mining_data.xlsx')

class Data_Excel:

    def __init__(self):
        self.setup_data_excel_file()

    def setup_data_excel_file(self) -> None:
        """
        Sets a file to write a data if not exist
        :return:
        """
        if os.path.isfile('mining_data.xlsx'):  # check if file exist
            return
        wb = openpyxl.Workbook()  # creating new file
        sheet = wb.active
        sheet['A1'] = "Date:"
        sheet['A2'] = "Current Hashrate:"
        sheet['A3'] = "Unpaid ETH:"
        sheet['A4'] = "Daily ETH:"
        sheet["A5"] = "Sum of payouts:"
        sheet["A6"] = "Days to next payout:"
        self.stretch_cells(sheet)  # stretch cells
        wb.save(filename="mining_data.xlsx")  # save file

    def stretch_cells(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Stretchs cells in first column
        :param sheet: openpyxl.worksheet.worksheet.Worksheet
        :return: None
        """
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value
    def checks_if_todays_data_is_saved(self) -> bool:
        """
        Checks if last saved data date is current returns True if yes and False if date is not
        :return: bool
        """
        wb = openpyxl.load_workbook('mining_data.xlsx')
        sheet = wb.active
        column_to_check = get_column_letter(sheet.max_column)
        if column_to_check == "A":  # checks if any data is saved
            return False
        date_to_check = sheet[column_to_check + "1"].value
        return date_to_check >= datetime(datetime.today().year, datetime.today().month, datetime.today().day)

if __name__ == "__main__":
    pass
    m = Miner(miner_address)
    # print(m.get_unpaid_eth())
    # print(m.get_sum_payouts())
    # print(m.get_daily_eth())
    # print(m.get_min_payment_threshold())
    # print(m.days_to_next_payout())
    # print(m.date_of_next_payout())
    # print(m.get_current_hashrate())
    # print(m.get_list_of_payouts())
    # m.save_todays_data_to_xlsx()
    de = Data_Excel()
    m.save_todays_data_to_xlsx()
