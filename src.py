from datetime import datetime, timedelta, date
from typing import List, Dict, TypeVar, Generic
import json
import os.path
from abc import ABC, abstractmethod

import requests
from pycoingecko import CoinGeckoAPI
import openpyxl
from openpyxl.utils import get_column_letter

from settings import miner_address, miner_cost, start_mining_date

Payout = TypeVar('Payout', bound=Dict[date, float])
List_Payouts = List[Payout]


class Miner:
    base_link = 'https://api.ethermine.org/miner/'
    cg = CoinGeckoAPI()
    # price ETH in USD and USD in PLN
    price_eth = round(float(cg.get_price(ids='ethereum', vs_currencies='usd')['ethereum']['usd']), 2)
    price_pln_as_usd = round(float(cg.get_price(ids='tether', vs_currencies='pln')['tether']['pln']), 2)

    def __init__(self, miner_address: str, miner_cost: str, start_mining_date: str, *args, **kwargs):
        self.miner_address = miner_address
        self.miner_cost = float(miner_cost)
        self.start_mining_date = start_mining_date

    @property
    def start_mining_date(self) -> date:
        return self.__start_mining_date

    @start_mining_date.setter
    def start_mining_date(self, start_mining_date: str):
        self.__start_mining_date = datetime.strptime(start_mining_date, '%d-%m-%Y')

    def get_unpaid_eth(self) -> float:
        """ Returns amount of unpaid_eth [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/dashboard")
        file = json.loads(res.text)
        return round(int(file['data']['currentStatistics']['unpaid']) / 10e17, 4)

    def get_sum_payouts(self) -> float:
        """ Returns sum of payouts [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/payouts")
        file = json.loads(res.text)
        return round(sum([int(elem['amount']) / 10e17 for elem in file['data']]), 4)

    def get_daily_eth(self) -> float:
        """ Returns daily mined eth [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/currentStats")
        file = json.loads(res.text)
        return round(float(file['data']['coinsPerMin']) * 24 * 60, 5)

    def get_min_payment_threshold(self) -> float:
        """ Returns minimum payment threshold [ETH] """

        res = requests.get(self.base_link + self.miner_address + "/settings")
        file = json.loads(res.text)
        return file['data']['minPayout'] / 10e17

    def days_to_next_payout(self) -> int:
        """ Returns days to next payout [Days] """

        return int((self.get_min_payment_threshold() - self.get_unpaid_eth()) / (self.get_daily_eth()))

    def date_of_next_payout(self) -> date:
        """ Returns date of  next payout [Y-M-D] """

        d = datetime.now() + timedelta(days=self.days_to_next_payout())
        return date(d.year, d.month, d.day)

    def get_list_of_payouts(self, order: str = "oldest") -> List_Payouts:
        """
        Returns a list of payout in form List[Dict[date, float]]
        if order == "oldest" list of payouts is sorted from first to last
        else from last to first
        :param order: str
        :return: List_Payouts
        """

        res = requests.get(self.base_link + self.miner_address + "/payouts")
        file = json.loads(res.text)
        list_of_payouts = []

        for payout in file['data']:
            d = datetime.fromtimestamp(payout['paidOn'])
            list_of_payouts.append({date(d.year, d.month, d.day): round(int(payout['amount']) / 10e17, 4)})
        if order == "oldest":
            list_of_payouts.reverse()
        return list_of_payouts

    def get_current_hashrate(self) -> float:
        """ Returns a current average hashrate of miner [MH/s] """

        res = requests.get(self.base_link + self.miner_address + "/dashboard")
        file = json.loads(res.text)
        return round(float(file['data']['statistics'][0]['reportedHashrate']) / 1000000, 1)

    def get_percentage_of_return_on_investment(self) -> float:
        """ Returns current percentage of return on investment as a sum of unpaid eth and sum of payouts [%] """
        return round((self.get_sum_payouts() + self.get_unpaid_eth()) * self.price_eth * self.price_pln_as_usd * 100
                     / self.miner_cost, 2)

    def get_total_duration_of_mining(self) -> int:
        """ Returns total duration of mining in days """
        d = datetime.now() - self.__start_mining_date
        return d.days


class Data_File(ABC):
    """ Abstract class that represents a data file """

    def __init__(self, *args, **kwargs):
        super(ABC, self).__init__(*args, **kwargs)

    @abstractmethod
    def setup_file(self):
        """ Setup a file to save data """
        raise NotImplementedError

    @abstractmethod
    def checks_if_todays_data_is_saved(self):
        """ Check if todays data is saved"""
        raise NotImplementedError


class Data_Excel(Data_File):
    """ Class represents data as excel file """

    def __init__(self, *args, **kwargs):
        super(Data_File, self).__init__(*args, **kwargs)
        self.setup_file()

    def setup_file(self) -> None:
        """ Sets a file to write a data if not exist """

        if os.path.isfile('mining_data.xlsx'):  # check if file exist
            return
        wb = openpyxl.Workbook()  # creating new file
        sheet = wb.active
        sheet['A1'] = "Date:"
        sheet['A2'] = "Current Hashrate:"
        sheet['A3'] = "Unpaid ETH:"
        sheet['A4'] = "Daily ETH:"
        sheet["A5"] = "Sum of payouts:"
        sheet["A6"] = "Days to next payout:"
        sheet["A7"] = "Today's ETH price [USD]:"
        self.stretch_cells(sheet)  # stretch cells
        wb.save(filename="mining_data.xlsx")  # save file

    def stretch_cells(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Stretchs cells in first column
        :param sheet: openpyxl.worksheet.worksheet.Worksheet
        :return: None
        """
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

    def checks_if_todays_data_is_saved(self) -> bool:
        """
        Checks if last saved data date is current returns True if yes and False if date is not
        :return: bool
        """
        wb = openpyxl.load_workbook('mining_data.xlsx')
        sheet = wb.active
        column_to_check = get_column_letter(sheet.max_column)
        if column_to_check == "A":  # checks if any data is saved
            return False
        date_to_check = sheet[column_to_check + "1"].value
        return date_to_check >= datetime(datetime.today().year, datetime.today().month, datetime.today().day)


class SaveData:
    """ Class allows to save data to files """

    def __init__(self, *args, **kwargs):
        self.miner = Miner(miner_address, miner_cost, start_mining_date)

    def save_todays_data_to_xlsx(self) -> None:
        """ If today's data didn't save -save it to xlsx file """

        de = Data_Excel()
        if de.checks_if_todays_data_is_saved():
            return
        wb = openpyxl.load_workbook('mining_data.xlsx')
        sheet = wb.active
        column_to_write = get_column_letter(sheet.max_column + 1)
        sheet[column_to_write + "1"] = date(datetime.today().year, datetime.today().month, datetime.today().day)
        sheet[column_to_write + "2"] = self.miner.get_current_hashrate()
        sheet[column_to_write + "3"] = self.miner.get_unpaid_eth()
        sheet[column_to_write + "4"] = self.miner.get_daily_eth()
        sheet[column_to_write + "5"] = self.miner.get_sum_payouts()
        sheet[column_to_write + "6"] = self.miner.days_to_next_payout()
        sheet[column_to_write + "7"] = self.miner.price_eth
        wb.save('mining_data.xlsx')

        wb = openpyxl.load_workbook('mining_data.xlsx')  # stretching cells
        de.stretch_cells(wb.active)
        wb.save('mining_data.xlsx')
